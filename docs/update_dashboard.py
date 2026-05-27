#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PT事業部 売上管理ダッシュボード 集計スクリプト
================================================
使い方：
    python update_dashboard.py --apo アポイントリスト.xlsx \
                               --prod 生産性レポート.csv \
                               --report レポート.csv \
                               --work 勤務データ.csv \
                               --inc インセンティブ.xlsx \  # 月次・任意
                               --master スタッフマスター.xlsx \
                               --prev pt_data.json \
                               --out pt_data_new.json

必須ファイル：
    --apo    : アポイントリスト（xlsx）
    --prod   : 生産性レポート（csv）
    --work   : 勤務データ（csv）
    --master : スタッフマスター（xlsx）
    --prev   : 前回のpt_data.json

任意ファイル：
    --report : レポート（csv）
    --inc    : インセンティブ（xlsx）※月1回

出力：
    --out    : 更新後のpt_data.json（デフォルト: pt_data_new.json）
"""

import argparse
import json
import re
import sys
from datetime import date, timedelta

import pandas as pd
from openpyxl import load_workbook

# ============================================================
# 定数・設定
# ============================================================
SITE_LABEL  = {'新宿SC': '新宿SC', '在宅G': 'リモートSC', 'AI': 'AI'}
B_TO_D      = {'B0000106': 'D0000295', 'B0000107': 'D0000326'}
KONO        = '幸野有希子CRM'   # 全体/サイト集計から除外
EXCLUDE_OPS = ['堀川璃歩']      # 全集計から除外
DAYS_IN_MONTH = 18              # 当月営業日数（設定で変更可能）


# ============================================================
# 営業日カレンダー
# ============================================================
def get_business_days(year: int, month: int, holidays: list = None) -> list:
    """指定年月の営業日一覧を返す"""
    if holidays is None:
        holidays = []
    biz = []
    d = date(year, month, 1)
    while d.month == month:
        if d.weekday() < 5 and d not in holidays:
            biz.append(d)
        d += timedelta(days=1)
    return biz


# ============================================================
# データ読み込み
# ============================================================
def load_staff_master(path: str) -> tuple:
    """スタッフマスター・時給マスターを読み込む"""
    wb = load_workbook(path, read_only=True)

    ws_m = wb['スタッフマスター']
    df_m = pd.DataFrame(
        list(ws_m.iter_rows(values_only=True))[1:],
        columns=['社員番号', 'スタッフ名', 'サイト', 'ランク']
    ).dropna(subset=['社員番号']).fillna('')
    df_m['社員番号'] = df_m['社員番号'].astype(str).str.strip()

    ws_w = wb['時給マスター']
    df_w = pd.DataFrame(
        list(ws_w.iter_rows(values_only=True))[1:],
        columns=['社員番号', 'スタッフ名', '時給', '備考']
    ).dropna(subset=['社員番号']).fillna('')
    df_w['社員番号'] = df_w['社員番号'].astype(str).str.strip()
    df_w['時給'] = pd.to_numeric(df_w['時給'], errors='coerce').fillna(0).astype(int)

    return df_m, df_w


def load_apo(path: str) -> pd.DataFrame:
    """アポイントリストを読み込む"""
    wb = load_workbook(path, read_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    # 除外OP
    df = df[~df['スタッフ名'].isin(EXCLUDE_OPS)].copy()
    # 君塚統合
    df['スタッフ名'] = df['スタッフ名'].replace('君塚綾子', '君塚綾子1104')
    df['cancel_date_str'] = df['キャンセル受付日'].astype(str).str.strip()
    df['sales'] = pd.to_numeric(df['案件金額'], errors='coerce').fillna(0)
    return df


def load_production(path: str) -> pd.DataFrame:
    """生産性レポートを読み込む"""
    for enc in ['utf-8-sig', 'cp932']:
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception:
            continue
    raise ValueError(f"生産性レポートの読み込みに失敗: {path}")


def load_work(path: str) -> pd.DataFrame:
    """勤務データを読み込む"""
    for enc in ['cp932', 'utf-8-sig']:
        try:
            df = pd.read_csv(path, encoding=enc)
            df = df.dropna(subset=['従業員ID'])
            df['総労働時間'] = pd.to_numeric(df['総労働時間'], errors='coerce').fillna(0)
            df['出勤日数']   = pd.to_numeric(df['出勤日数'],   errors='coerce').fillna(0)
            # 重複グループ：最大労働時間採用
            df = df.sort_values('総労働時間', ascending=False).drop_duplicates(subset='従業員ID')
            return df
        except Exception:
            continue
    raise ValueError(f"勤務データの読み込みに失敗: {path}")


def load_incentive(path: str) -> dict:
    """インセンティブファイルを読み込む"""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    result = {}
    for row in rows:
        if len(row) >= 3 and row[2]:
            name = str(row[1]).strip() if row[1] else ''
            amount = int(pd.to_numeric(row[2], errors='coerce') or 0)
            if name and amount > 0:
                result[name] = amount
    return result


# ============================================================
# 集計処理
# ============================================================
def calc_labor(df_work: pd.DataFrame, df_master: pd.DataFrame,
               df_wage: pd.DataFrame, days_in_month: int) -> tuple:
    """人件費を計算する。(site_labor_dict, work_by_id, labor_by_id) を返す"""
    master_ids  = set(df_master['社員番号'].tolist())
    site_map_id = dict(zip(df_master['社員番号'], df_master['サイト']))
    wage_by_id  = dict(zip(df_wage['社員番号'], df_wage['時給']))
    note_by_id  = dict(zip(df_wage['社員番号'], df_wage['備考']))

    df_active = df_work[(df_work['出勤日数'] >= 1) & (df_work['総労働時間'] > 0)].copy()

    site_labor = {'新宿SC': 0, 'リモートSC': 0, 'AI': 0}
    work_by_id = {}
    labor_by_id = {}

    for _, row in df_active.iterrows():
        emp_id = str(row['従業員ID']).strip()
        hours  = float(row['総労働時間'])
        days   = float(row['出勤日数'])
        lookup = B_TO_D.get(emp_id, emp_id)

        if lookup not in master_ids and emp_id not in master_ids:
            continue

        site_r = site_map_id.get(lookup) or site_map_id.get(emp_id, '')
        site_d = SITE_LABEL.get(site_r, 'その他')
        wage   = wage_by_id.get(lookup) or wage_by_id.get(emp_id)
        note   = str(note_by_id.get(lookup) or note_by_id.get(emp_id, ''))

        if not wage:
            continue

        wage = float(wage)
        cost = (round(wage * 1.15 / days_in_month * days)
                if '月給' in note else round(wage * hours))

        site_labor[site_d] = site_labor.get(site_d, 0) + cost
        work_by_id[lookup] = work_by_id[emp_id] = hours
        labor_by_id[lookup] = labor_by_id[emp_id] = cost

    return site_labor, work_by_id, labor_by_id


def aggregate_apo(df_apo: pd.DataFrame, target_dates: list,
                  site_map: dict) -> tuple:
    """
    アポイントリストから集計する。
    returns: (op_get, op_cxl_dash, op_cxl_op)
    """
    # 獲得（再送除外・幸野除外）
    df_get = df_apo[
        df_apo['取得日'].isin(target_dates) &
        (df_apo['再送当否'].astype(str).str.strip() != '再送') &
        (df_apo['スタッフ名'] != KONO)
    ].copy()
    df_get['site_raw'] = df_get['スタッフ名'].map(site_map)

    # キャンセル（ダッシュボード用・考慮除外）
    df_cxl_dash = df_apo[
        df_apo['cancel_date_str'].isin(target_dates) &
        (df_apo['スタッフ名'] != KONO)
    ].copy()
    df_cxl_dash['site_raw'] = df_cxl_dash['スタッフ名'].map(site_map)

    # キャンセル（OP実績用・考慮込み）
    kouryo_mask = (
        df_apo['cancel_date_str'].str.contains('考慮', na=False) &
        df_apo['cancel_date_str'].str.extract(r'(2026/\d+/\d+)', expand=False)
              .isin(target_dates)
    )
    df_cxl_op = df_apo[
        df_apo['cancel_date_str'].isin(target_dates) | kouryo_mask
    ].copy()

    return df_get, df_cxl_dash, df_cxl_op


def calc_daily(df_apo: pd.DataFrame, date_str: str,
               site_map: dict, calls_by_date: dict,
               ops_by_date: dict) -> dict:
    """1営業日分の集計を返す"""
    df_g = df_apo[
        (df_apo['取得日'] == date_str) &
        (df_apo['再送当否'].astype(str).str.strip() != '再送') &
        (df_apo['スタッフ名'] != KONO)
    ].copy()
    df_g['site_raw'] = df_g['スタッフ名'].map(site_map)

    df_c = df_apo[
        (df_apo['cancel_date_str'] == date_str) &
        (df_apo['スタッフ名'] != KONO)
    ].copy()
    df_c['site_raw'] = df_c['スタッフ名'].map(site_map)

    result = {}
    for key, raw in [('all', None), ('shinjuku', '新宿SC'),
                     ('remote', '在宅G'), ('ai', 'AI')]:
        g = df_g if raw is None else df_g[df_g['site_raw'] == raw]
        c = df_c if raw is None else df_c[df_c['site_raw'] == raw]
        gs = int(g['sales'].sum())
        cs = int(c['sales'].sum())
        result[key] = {
            'sales':  gs - cs,
            'apo':    len(g),
            'cancel': len(c),
            'valid':  len(g) - len(c),
            'calls':  calls_by_date.get(date_str, 0) if key == 'all' else 0,
            'ops':    ops_by_date.get(date_str, 0)
        }
    return result


def calc_heatmap(df_apo: pd.DataFrame, target_dates: list,
                 total_sales: int) -> list:
    """ヒートマップTOP10を計算する"""
    df_g = df_apo[
        df_apo['取得日'].isin(target_dates) &
        (df_apo['再送当否'].astype(str).str.strip() != '再送') &
        (df_apo['スタッフ名'] != KONO)
    ].copy()
    df_c = df_apo[
        df_apo['cancel_date_str'].isin(target_dates) &
        (df_apo['スタッフ名'] != KONO)
    ].copy()

    pg = df_g.groupby('登録案件名').agg(
        apo=('アポイントID', 'count'), sg=('sales', 'sum')).reset_index()
    pc = df_c.groupby('登録案件名').agg(
        cxl=('アポイントID', 'count'), sc=('sales', 'sum')).reset_index()
    proj = pg.merge(pc, on='登録案件名', how='left').fillna(0)
    proj['valid'] = (proj['apo'] - proj['cxl']).astype(int)
    proj['net']   = (proj['sg']  - proj['sc']).astype(int)
    proj = proj[proj['valid'] > 0].sort_values('net', ascending=False).head(10)

    return [
        {
            'rank':  i + 1,
            'name':  str(r['登録案件名']),
            'valid': int(r['valid']),
            'sales': int(r['net']),
            'pct':   round(r['net'] / total_sales * 100, 1) if total_sales > 0 else 0
        }
        for i, (_, r) in enumerate(proj.iterrows())
    ]


def calc_operators(df_apo: pd.DataFrame, target_dates: list,
                   df_master: pd.DataFrame, df_prod: pd.DataFrame,
                   work_by_id: dict, labor_by_id: dict,
                   inc_map: dict, elapsed: int, working: int,
                   prev_calls: dict) -> list:
    """OP個人実績を計算する"""
    site_map  = dict(zip(df_master['スタッフ名'], df_master['サイト']))
    rank_map  = dict(zip(df_master['スタッフ名'], df_master['ランク']))
    id_map    = dict(zip(df_master['スタッフ名'], df_master['社員番号']))

    df_get = df_apo[
        df_apo['取得日'].isin(target_dates) &
        (df_apo['再送当否'].astype(str).str.strip() != '再送')
    ].copy()
    kouryo_mask = (
        df_apo['cancel_date_str'].str.contains('考慮', na=False) &
        df_apo['cancel_date_str'].str.extract(r'(2026/\d+/\d+)', expand=False)
              .isin(target_dates)
    )
    df_cxl_op   = df_apo[df_apo['cancel_date_str'].isin(target_dates) | kouryo_mask].copy()
    df_cxl_dash = df_apo[df_apo['cancel_date_str'].isin(target_dates)].copy()

    op_get    = df_get.groupby('スタッフ名').agg(
        apo=('アポイントID', 'count'), sg=('sales', 'sum')).reset_index()
    op_cxl_op = df_cxl_op.groupby('スタッフ名').agg(
        cxl_op=('アポイントID', 'count'), sc_op=('sales', 'sum')).reset_index()
    op_cxl_d  = df_cxl_dash.groupby('スタッフ名').agg(
        cxl_d=('アポイントID', 'count'), sc_d=('sales', 'sum')).reset_index()

    calls_new = df_prod.groupby('エージェント')['コール数'].sum().to_dict()
    all_names = set(list(df_master['スタッフ名']) + list(op_get['スタッフ名']))

    operators = []
    for name in all_names:
        g  = op_get[op_get['スタッフ名'] == name]
        co = op_cxl_op[op_cxl_op['スタッフ名'] == name]
        cd = op_cxl_d[op_cxl_d['スタッフ名'] == name]

        apo    = int(g['apo'].iloc[0])     if len(g)  > 0 else 0
        sg     = int(g['sg'].iloc[0])      if len(g)  > 0 else 0
        cxl_op = int(co['cxl_op'].iloc[0]) if len(co) > 0 else 0
        sc_op  = int(co['sc_op'].iloc[0])  if len(co) > 0 else 0
        cxl_d  = int(cd['cxl_d'].iloc[0]) if len(cd) > 0 else 0
        sc_d   = int(cd['sc_d'].iloc[0])  if len(cd) > 0 else 0

        net_op   = sg - sc_op
        net_dash = sg - sc_d
        calls    = prev_calls.get(name, 0) + int(calls_new.get(name, 0))

        emp_id  = id_map.get(name, '')
        lookup  = B_TO_D.get(emp_id, emp_id)
        hours   = work_by_id.get(lookup) or work_by_id.get(emp_id, 0)
        l_base  = labor_by_id.get(lookup) or labor_by_id.get(emp_id, 0)
        inc_t   = inc_map.get(name, 0)
        inc_day = round(inc_t / working * elapsed) if inc_t > 0 else 0
        labor   = l_base + inc_day

        days    = float(hours / (labor / l_base)) if l_base > 0 and hours > 0 else 0
        # 出勤日数は勤務データから取得（簡易的にlabor_baseから逆算せず別途渡す方が正確）
        unit_per_day = round(net_op / days) if days > 0 and net_op > 0 else 0

        ar     = round(apo / calls * 100, 1) if calls > 0 else None
        cost_r = round(labor / net_op * 100, 1) if net_op > 0 and labor > 0 else None

        operators.append({
            'name':            name,
            'site':            SITE_LABEL.get(site_map.get(name, ''), ''),
            'rank':            rank_map.get(name, ''),
            'sales':           net_op,
            'sales_dash':      net_dash,
            'apo':             apo,
            'cancel':          cxl_op,
            'valid':           apo - cxl_op,
            'calls':           calls,
            'apoRate':         ar,
            'workH':           round(float(hours), 1),
            'labor':           labor,
            'labor_base':      l_base,
            'incentive_daily': inc_day,
            'days':            days,
            'unitPerDay':      unit_per_day,
            'costRate':        cost_r,
        })

    operators.sort(key=lambda x: (-x['sales'] if x['sales'] > 0
                                  else (0 if x['sales'] == 0 else 1)))
    return operators


# ============================================================
# メイン処理
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='PT事業部ダッシュボード集計スクリプト')
    parser.add_argument('--apo',    required=True,  help='アポイントリスト xlsx')
    parser.add_argument('--prod',   required=True,  help='生産性レポート csv')
    parser.add_argument('--report', required=False, help='レポート csv（任意）')
    parser.add_argument('--work',   required=True,  help='勤務データ csv')
    parser.add_argument('--master', required=True,  help='スタッフマスター xlsx')
    parser.add_argument('--prev',   required=True,  help='前回の pt_data.json')
    parser.add_argument('--inc',    required=False, help='インセンティブ xlsx（月次・任意）')
    parser.add_argument('--out',    default='pt_data_new.json', help='出力先 json')
    args = parser.parse_args()

    print("=== PT事業部ダッシュボード集計スクリプト ===")

    # --- データ読み込み ---
    print("データ読み込み中...")
    df_master, df_wage = load_staff_master(args.master)
    df_apo  = load_apo(args.apo)
    df_prod = load_production(args.prod)
    df_work = load_work(args.work)

    with open(args.prev, encoding='utf-8') as f:
        PT_DATA = json.load(f)

    inc_map = PT_DATA.get('incentive', {})
    if args.inc:
        print("インセンティブデータを読み込みます...")
        inc_map = load_incentive(args.inc)

    # --- 基本設定 ---
    elapsed  = PT_DATA['meta']['elapsedDays']
    working  = PT_DATA['meta']['workingDays']
    site_map = dict(zip(df_master['スタッフ名'], df_master['サイト']))

    # --- 取得日の特定 ---
    existing_dates = set(
        row['date'] for rows in PT_DATA['daily'].values() for row in rows
    )
    all_dates_in_apo = set(df_apo['取得日'].dropna().unique())
    new_dates = sorted(all_dates_in_apo - existing_dates)

    if not new_dates:
        print("新しい営業日データが見つかりません。終了します。")
        sys.exit(0)

    print(f"新規営業日: {new_dates}")
    target_dates = sorted(existing_dates | set(new_dates))

    # --- 人件費計算 ---
    print("人件費計算中...")
    site_labor, work_by_id, labor_by_id = calc_labor(
        df_work, df_master, df_wage, working)

    # --- コール・稼働人数 ---
    calls_by_date = {}
    ops_by_date   = {}
    for d in new_dates:
        d_key = d.replace('/', '-').replace('2026-', '2026-')
        # 生産性レポートの日付形式に合わせる
        d_prod = d.replace('/', '-') if '-' in str(df_prod['日付'].iloc[0]) \
                 else d
        mask = df_prod['日付'] == d_prod
        calls_by_date[d] = int(df_prod[mask]['コール数'].sum())
        ops_by_date[d]   = len(df_prod[mask])

    # --- 日次明細追加 ---
    print("日次明細を追加中...")
    for i, date_str in enumerate(new_dates):
        elapsed += 1
        day_num  = elapsed
        daily    = calc_daily(df_apo, date_str, site_map,
                              calls_by_date, ops_by_date)
        for key in ['all', 'shinjuku', 'remote', 'ai']:
            d = daily[key]
            PT_DATA['daily'][key].append({
                'day':    f'{day_num}営業日',
                'date':   date_str,
                'sales':  d['sales'],
                'apo':    d['apo'],
                'cancel': d['cancel'],
                'valid':  d['valid'],
                'ops':    d['ops'],
                'calls':  d['calls'],
            })

    PT_DATA['meta']['elapsedDays'] = elapsed

    # --- サイト別累計 ---
    print("サイト別累計を計算中...")
    inc_site = {'新宿SC': 0, 'リモートSC': 0, 'AI': 0}
    for name, inc_total in inc_map.items():
        if inc_total <= 0:
            continue
        site = SITE_LABEL.get(site_map.get(name, ''), '')
        if site in inc_site:
            inc_site[site] += round(inc_total / working * elapsed)
    inc_all = sum(inc_site.values())

    for k in ['all', 'shinjuku', 'remote', 'ai']:
        rows    = PT_DATA['daily'][k]
        sales   = sum(r['sales']  for r in rows)
        apo     = sum(r['apo']    for r in rows)
        cancel  = sum(r['cancel'] for r in rows)
        calls   = sum(r['calls']  for r in rows)
        valid   = apo - cancel
        cr      = round(cancel / apo * 100, 1) if apo > 0 else 0
        site_jp = {'all': None, 'shinjuku': '新宿SC',
                   'remote': 'リモートSC', 'ai': 'AI'}[k]
        jinjer  = {'all': sum(site_labor.values()),
                   'shinjuku': site_labor.get('新宿SC', 0),
                   'remote':   site_labor.get('リモートSC', 0),
                   'ai':       site_labor.get('AI', 0)}[k]
        labor   = jinjer + (inc_all if k == 'all'
                            else inc_site.get(site_jp, 0))
        cost    = round(labor / sales * 100, 1) if sales > 0 else 0
        ar      = round(apo / calls * 100, 2)   if calls > 0 else 0
        last    = rows[-1] if rows else {}
        unit    = (round(last['sales'] / last['ops'])
                   if last and last.get('ops', 0) > 0 else 0)

        PT_DATA['sites'][k] = {
            'sales': sales, 'apo': apo, 'cancel': cancel, 'valid': valid,
            'cancelRate': cr, 'labor': labor, 'costRate': cost,
            'gross': sales - labor, 'ops': last.get('ops', 0),
            'unit': unit, 'calls': calls, 'apoRate': ar,
        }

    # --- チャート ---
    print("チャートデータを更新中...")
    chart_labels = PT_DATA['chart']['labels']
    chart_data   = [0.0] * len(chart_labels)
    for row in PT_DATA['daily']['all']:
        d_label = row['date'].replace('2026/', '').lstrip('0').replace('/0', '/')
        if d_label in chart_labels:
            idx = chart_labels.index(d_label)
            chart_data[idx] = round(row['sales'] / 10000, 1)
    PT_DATA['chart']['data'] = chart_data

    # --- ヒートマップ ---
    print("ヒートマップを更新中...")
    PT_DATA['heatmap'] = calc_heatmap(
        df_apo, list(target_dates), PT_DATA['sites']['all']['sales'])

    # --- OP個人実績 ---
    print("OP個人実績を計算中...")
    prev_calls = PT_DATA.get('prev_calls', {})
    PT_DATA['operators'] = calc_operators(
        df_apo, list(target_dates), df_master, df_prod,
        work_by_id, labor_by_id, inc_map,
        elapsed, working, prev_calls)

    # --- メタ更新 ---
    today = date.today().strftime('%Y/%m/%d')
    PT_DATA['meta']['lastUpdate']      = today
    PT_DATA['meta']['lastUpdateLabel'] = (
        f"{date.today().strftime('%m/%d')}（{new_dates[-1].replace('2026/', '')}分反映済）")
    PT_DATA['meta']['alertText'] = (
        f"{new_dates[-1].replace('2026/', '')}のデータを反映済みです"
        f"（累計{elapsed}営業日）。最終更新: {today}")

    if args.inc:
        PT_DATA['incentive'] = inc_map

    # --- 検証 ---
    j = json.dumps(PT_DATA, ensure_ascii=False)
    assert 'NaN' not in j,      "エラー: NaNが含まれています"
    assert 'Infinity' not in j, "エラー: Infinityが含まれています"

    # --- 出力 ---
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(PT_DATA, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 完了: {args.out} に出力しました")
    print(f"   経過営業日数: {elapsed}日")
    print(f"   累計売上: ¥{PT_DATA['sites']['all']['sales']:,}")
    print(f"   原価率: {PT_DATA['sites']['all']['costRate']}%")


if __name__ == '__main__':
    main()
